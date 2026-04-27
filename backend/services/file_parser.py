import io
import os
from typing import Optional


def extract_pdf_text(file_path: str) -> str:
    """Extract text from a PDF file."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        texts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                texts.append(text)
        return "\n\n".join(texts) if texts else "[PDF 内容为空或无法提取文本]"
    except ImportError:
        return "[PDF 解析依赖未安装: pypdf]"
    except Exception as e:
        return f"[PDF 解析失败: {e}]"


def extract_excel_text(file_path: str) -> str:
    """Extract text from an Excel file as markdown table."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}")
            rows = []
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                rows.append(row_text)
            if rows:
                parts.append("\n".join(rows))
        return "\n\n".join(parts) if parts else "[Excel 内容为空]"
    except ImportError:
        return "[Excel 解析依赖未安装: openpyxl]"
    except Exception as e:
        return f"[Excel 解析失败: {e}]"


def get_file_text(file_path: str, file_type: Optional[str]) -> str:
    """Extract text content from a file based on its type."""
    # Convert /uploads/ URL paths to local filesystem paths
    if file_path.startswith("/uploads/"):
        from pathlib import Path
        backend_dir = Path(__file__).parent.parent
        file_path = str(backend_dir / "uploads" / file_path[len("/uploads/"):])
    if not os.path.exists(file_path):
        return "[文件不存在]"
    if file_type == "pdf":
        return extract_pdf_text(file_path)
    if file_type == "excel":
        return extract_excel_text(file_path)
    if file_type == "image":
        return "[图片文件，无法直接提取文本内容]"
    return "[不支持的文件类型]"
